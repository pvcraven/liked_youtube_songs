"""
This program will take your
"""
import json
from dataclasses import dataclass
from typing import List

from ytmusicapi import YTMusic
from openpyxl import Workbook
import re
from openpyxl import utils


@dataclass
class Song:
    title: str
    artists: List
    hyperlink: str


def as_text(value):
    if value is None:
        return ""
    return str(value)


def main():
    """ Main method where the magic happens. """
    try:
        print("Authorizing...")
        ytmusic = YTMusic("oauth.json")
    except json.decoder.JSONDecodeError:
        print("Error reading oauth.json file. Run 'ytmusicapi oauth' from the command line to generate.")
        return

    print("Getting liked songs...")
    song_query_result = ytmusic.get_liked_songs(2000)

    song_list = []
    possible_artist_list = []
    with open('valid_artists.txt') as f:
        valid_artist_list = f.read().splitlines()
    title_re = re.compile('(^[^\(-]*) - (.*)$')
    for track in song_query_result['tracks']:
        artist_list = []
        for artists in track['artists']:
            artist_list.append(artists['name'])
        title = track['title']
        re_match = title_re.match(title)
        if re_match and re_match.group(1).lower() == artist_list[0].lower():
            # Fix if the title has the artist, which is the same as the artist that's already listed.
            title = re_match.group(2)
            print(f"Removed duplicate artist {artist_list[0]} from title.")
        elif re_match and re_match.group(1) in valid_artist_list:
            title = re_match.group(2)
            artist_list.insert(0, re_match.group(1))
            print(f"Fix for artist {re_match.group(1)}")
        elif re_match:
            print(title)
            print(f"Original Artist: {artist_list}")
            print(f"Artist:          {re_match.group(1)}")
            print(f"Song:            {re_match.group(2)}")
            possible_artist_list.append(re_match.group(1))
            print()

        song = Song(title=title, artists=artist_list, hyperlink=track['videoId'])
        song_list.append(song)

    sorted_song_list = sorted(song_list, key=lambda x: x.artists[0].lower() + x.title)
    # print(songs)
    # print(json.dumps(songs, indent=4))
    print("Writing out spreadsheet...")

    workbook = Workbook()
    work_sheet = workbook.active
    work_sheet.append(['Title', 'Artist 1', 'Artist 2', 'Artist 3'])

    for song in sorted_song_list:
        row = [song.title]
        for artist in song.artists:
            row.append(artist)

        work_sheet.append(row)
        last_row = len(work_sheet['A'])
        cell = work_sheet.cell(last_row, 1)
        cell.hyperlink = f"https://music.youtube.com/watch?v={song.hyperlink}"

    # Set column width
    for column_cells in work_sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        work_sheet.column_dimensions[utils.get_column_letter(column_cells[0].column)].width = length

    workbook.save("liked_songs.xlsx")
    print("Finished. Results in liked_songs.xlsx")

    print(possible_artist_list)
    possible_artist_list = set(possible_artist_list)
    possible_artist_list = list(possible_artist_list)
    possible_artist_list.sort()
    for artist in possible_artist_list:
        print(artist)


main()
